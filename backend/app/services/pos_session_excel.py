from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.document_telegram_format import (
    _format_date,
    _item_name,
    _payment_type_name,
    _visible_pos_cheque_operations,
)
from app.services.pos_session_report import SessionReportData
from app.services.telegram_i18n import t

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

_DATA_FONT = Font(size=11)
_TEXT_ALIGNMENT = Alignment(vertical="center")
_NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")
_CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

_ALT_ROW_FILL = PatternFill("solid", fgColor="F5F7FA")
_RETURN_ROW_FILL = PatternFill("solid", fgColor="FCE4D6")
_THIN_SIDE = Side(style="thin", color="D0D5DD")
_CELL_BORDER = Border(
    left=_THIN_SIDE,
    right=_THIN_SIDE,
    top=_THIN_SIDE,
    bottom=_THIN_SIDE,
)

_NUMBER_FORMAT = "#,##0.00"
_QUANTITY_FORMAT = "#,##0.00"


def _cheque_uuid(cheque: dict[str, Any]) -> str:
    return str(cheque.get("uuid", "")).strip().lower()


def _sale_return_label(is_return: bool, lang: str) -> str:
    key = (
        "telegram.receipt.posSessionExcelReturn"
        if is_return
        else "telegram.receipt.posSessionExcelSale"
    )
    return t(key, lang)


def _autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 40)


def _style_header_row(sheet, *, column_count: int) -> None:
    for column_index in range(1, column_count + 1):
        cell = sheet.cell(row=1, column=column_index)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _CELL_BORDER
    sheet.row_dimensions[1].height = 24


def _style_data_sheet(
    sheet,
    *,
    column_count: int,
    numeric_columns: set[int],
    quantity_columns: set[int] | None = None,
    return_column: int | None = None,
    return_label: str | None = None,
) -> None:
    quantity_columns = quantity_columns or set()

    for row_index in range(2, sheet.max_row + 1):
        is_return_row = False
        if return_column is not None and return_label is not None:
            cell_value = sheet.cell(row=row_index, column=return_column).value
            is_return_row = cell_value == return_label

        use_alt_fill = row_index % 2 == 0
        row_fill = _RETURN_ROW_FILL if is_return_row else (_ALT_ROW_FILL if use_alt_fill else None)

        for column_index in range(1, column_count + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.font = _DATA_FONT
            cell.border = _CELL_BORDER

            if row_fill is not None:
                cell.fill = row_fill

            if column_index in numeric_columns:
                cell.alignment = _NUMBER_ALIGNMENT
                cell.number_format = (
                    _QUANTITY_FORMAT
                    if column_index in quantity_columns
                    else _NUMBER_FORMAT
                )
            elif return_column is not None and column_index == return_column:
                cell.alignment = _CENTER_ALIGNMENT
            else:
                cell.alignment = _TEXT_ALIGNMENT

    if sheet.max_row >= 1:
        last_column = get_column_letter(column_count)
        sheet.auto_filter.ref = f"A1:{last_column}{sheet.max_row}"
    sheet.freeze_panes = "A2"


def session_report_filename(cash_session: dict[str, Any]) -> str:
    session_code = str(cash_session.get("code") or "session").strip() or "session"
    return f"{session_code}-report.xlsx"


def generate_session_excel(report: SessionReportData, *, lang: str = "ru") -> bytes:
    workbook = Workbook()
    return_label = t("telegram.receipt.posSessionExcelReturn", lang)

    products_sheet = workbook.active
    products_sheet.title = t("telegram.receipt.posSessionExcelProductsSheet", lang)
    products_sheet.append(
        [
            t("common.date", lang),
            t("telegram.receipt.posSessionExcelCheque", lang),
            t("telegram.receipt.items", lang),
            t("telegram.receipt.posSessionExcelQuantity", lang),
            t("telegram.receipt.posSessionExcelPrice", lang),
            t("common.amount", lang),
            t("telegram.receipt.posSessionExcelSaleReturn", lang),
        ]
    )

    for cheque in report.cheques:
        cheque_uuid = _cheque_uuid(cheque)
        cheque_code = cheque.get("code", "")
        cheque_date = _format_date(cheque.get("date"))
        is_return = bool(cheque.get("is_return"))
        sale_return = _sale_return_label(is_return, lang)
        operations = _visible_pos_cheque_operations(
            report.operations_by_cheque.get(cheque_uuid, [])
        )

        for operation in operations:
            if bool(operation.get("has_storno")):
                continue
            quantity = float(operation.get("quantity", 0))
            price = float(operation.get("price", 0))
            products_sheet.append(
                [
                    cheque_date,
                    cheque_code,
                    _item_name(operation.get("item"), lang),
                    quantity,
                    price,
                    quantity * price,
                    sale_return,
                ]
            )

    products_column_count = 7
    _style_header_row(products_sheet, column_count=products_column_count)
    _style_data_sheet(
        products_sheet,
        column_count=products_column_count,
        numeric_columns={4, 5, 6},
        quantity_columns={4},
        return_column=7,
        return_label=return_label,
    )
    _autosize_columns(products_sheet)

    payments_sheet = workbook.create_sheet(
        title=t("telegram.receipt.posSessionExcelPaymentsSheet", lang)
    )
    payments_sheet.append(
        [
            t("common.date", lang),
            t("telegram.receipt.posSessionExcelCheque", lang),
            t("common.type", lang),
            t("common.amount", lang),
            t("telegram.receipt.posSessionExcelSaleReturn", lang),
        ]
    )

    for cheque in report.cheques:
        cheque_uuid = _cheque_uuid(cheque)
        cheque_code = cheque.get("code", "")
        cheque_date = _format_date(cheque.get("date"))
        is_return = bool(cheque.get("is_return"))
        sale_return = _sale_return_label(is_return, lang)

        for payment in report.payments_by_cheque.get(cheque_uuid, []):
            if bool(payment.get("has_storno")):
                continue
            value = float(payment.get("value", 0))
            if value < 0:
                continue
            payments_sheet.append(
                [
                    cheque_date,
                    cheque_code,
                    _payment_type_name(payment, lang),
                    value,
                    sale_return,
                ]
            )

    payments_column_count = 5
    _style_header_row(payments_sheet, column_count=payments_column_count)
    _style_data_sheet(
        payments_sheet,
        column_count=payments_column_count,
        numeric_columns={4},
        return_column=5,
        return_label=return_label,
    )
    _autosize_columns(payments_sheet)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
