from __future__ import annotations

import io
from datetime import UTC, datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.services.telegram_i18n import t

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

_DATA_FONT = Font(size=11)
_TEXT_ALIGNMENT = Alignment(vertical="center")
_NUMBER_ALIGNMENT = Alignment(horizontal="right", vertical="center")

_ALT_ROW_FILL = PatternFill("solid", fgColor="F5F7FA")
_THIN_SIDE = Side(style="thin", color="D0D5DD")
_CELL_BORDER = Border(
    left=_THIN_SIDE,
    right=_THIN_SIDE,
    top=_THIN_SIDE,
    bottom=_THIN_SIDE,
)

_NUMBER_FORMAT = "#,##0.00"
_QUANTITY_FORMAT = "#,##0.00"


def out_of_stock_report_filename() -> str:
    return "out-of-stock-report.xlsx"


def _format_detected_at(value: Any) -> str:
    if isinstance(value, datetime):
        dt = value.astimezone(UTC).replace(tzinfo=None) if value.tzinfo else value
        return dt.strftime("%d.%m.%Y %H:%M")
    if isinstance(value, (int, float)):
        try:
            return datetime.fromtimestamp(value).strftime("%d.%m.%Y %H:%M")
        except (OSError, OverflowError, ValueError):
            return str(value)
    return str(value) if value else ""


def _column_headers(lang: str) -> list[str]:
    return [
        t("dashboard.outOfStock.columns.product", lang),
        t("dashboard.outOfStock.columns.code", lang),
        t("dashboard.outOfStock.columns.barcode", lang),
        t("dashboard.outOfStock.columns.warehouse", lang),
        t("dashboard.outOfStock.columns.quantity", lang),
        t("dashboard.outOfStock.columns.minQuantity", lang),
        t("dashboard.outOfStock.columns.lastPurchaseCost", lang),
        t("dashboard.outOfStock.columns.price", lang),
        t("dashboard.outOfStock.columns.detectedAt", lang),
    ]


def _autosize_columns(sheet) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_length = max(max_length, len(str(value)))
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 40)


def _style_header_row(sheet, *, column_count: int) -> None:
    for column_index in range(1, column_count + 1):
        cell = sheet.cell(row=1, column=column_index)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _CELL_BORDER
    sheet.row_dimensions[1].height = 24


def _style_data_sheet(
    sheet,
    *,
    column_count: int,
    numeric_columns: set[int],
    quantity_columns: set[int] | None = None,
) -> None:
    quantity_columns = quantity_columns or set()

    for row_index in range(2, sheet.max_row + 1):
        use_alt_fill = row_index % 2 == 0
        row_fill = _ALT_ROW_FILL if use_alt_fill else None

        for column_index in range(1, column_count + 1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.font = _DATA_FONT
            cell.border = _CELL_BORDER

            if row_fill is not None:
                cell.fill = row_fill

            if column_index in numeric_columns:
                cell.alignment = _NUMBER_ALIGNMENT
                cell.number_format = (
                    _QUANTITY_FORMAT
                    if column_index in quantity_columns
                    else _NUMBER_FORMAT
                )
            else:
                cell.alignment = _TEXT_ALIGNMENT

    if sheet.max_row >= 1:
        last_column = get_column_letter(column_count)
        sheet.auto_filter.ref = f"A1:{last_column}{sheet.max_row}"
    sheet.freeze_panes = "A2"


def generate_out_of_stock_excel(rows: list[dict[str, Any]], *, lang: str = "ru") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = t("dashboard.outOfStock.title", lang)[:31]
    sheet.append(_column_headers(lang))

    for row in rows:
        last_purchase_cost = row.get("last_purchase_cost")
        sheet.append(
            [
                row.get("product_name", ""),
                row.get("code", ""),
                row.get("barcode", ""),
                row.get("stock_name", ""),
                float(row.get("quantity", 0)),
                float(row.get("min_quantity", 0)),
                float(last_purchase_cost) if last_purchase_cost is not None else None,
                float(row.get("price", 0)),
                _format_detected_at(row.get("detected_at")),
            ]
        )

    column_count = 9
    _style_header_row(sheet, column_count=column_count)
    _style_data_sheet(
        sheet,
        column_count=column_count,
        numeric_columns={5, 6, 7, 8},
        quantity_columns={5, 6},
    )
    _autosize_columns(sheet)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
